from bs4 import BeautifulSoup
import requests, openpyxl, os, re, openpyxl.styles, random
from openpyxl.styles import Alignment, Font


URL_HOME = "https://books.toscrape.com/"

def getAllCategoriesTitles(url:str)-> dict:
    """
    Cette fonction permet de récuperer les titres et les liens de chaque catégories de la page d'accueil.
    Elle prend en paramètre l'url de la page d'acceuil.
    Elle retourne un dictionnaire avec les catégories comme clefs et les urls des catégories comme valeurs

    """ 

    CONTENT_BODY_HOME = requests.get(url).text
    reponse = requests.get(url)

    soup = BeautifulSoup(CONTENT_BODY_HOME, "html.parser")

    # récupération du menu "catégories"
    menu_nav_categories = soup.find("ul", class_="nav nav-list")
    ul_category = menu_nav_categories.li.ul

    # récuperation des titres
    tab_titles = []
    category_title = ul_category.li.a
    for category_title in ul_category:
        category_title_strip = category_title.getText('',True)
        if not category_title_strip == "":
            tab_titles.append(category_title_strip)

    #récupération des liens
    tab_links = []
    all_tag_a = ul_category.find_all('a')
    for tag_a in all_tag_a:
        link = url+tag_a['href']
        tab_links.append(link)

    # création du dictionnaire
    dictionnary = dict(zip(tab_titles[0:2],tab_links[0:2]))

    return dictionnary



def create_excel_files_by_categories_names(folder_categories_name: str, dictionnary_categories_names_and_url: dict)->bool:
    """
    Cette fonction permet de créer un dossier contenant des fichiers excels. 
    Elle prend 2 paramètres :
        - le nom du répertoire à créer
        - le dictionnaire dont les clefs seront les noms des fichiers excels à créer

    """
    nb_categories = len(dictionnary_categories_names_and_url)
    counter = 0
    for categorie_name in dictionnary_categories_names_and_url.keys():
        file = openpyxl.Workbook()
        file.save(f"{folder_categories_name}/{categorie_name}.xlsx")
        counter += 1
        print(f"Création du fichier {categorie_name}. {counter}/{nb_categories}")

    if counter == nb_categories:
        return True
    else:
        return False



def is_many_pages(soup:object)->bool:
    """
    Fonction permettant de savoir si le bouton "next" est présent une page.
    Elle prend en paramètre un objet "soup".
    Elle retourne un booléen.

    """
    li_next = soup.find("li", class_="next")
    return li_next



# tant qu'il y a de h3 (livres) on places les liens et les titles de la balise "a" dans le dictionnaire

def get_links_articles_for_one_categorie(url_article, tab, numero_page=1):
    """
    Cette fonction permets de récuperer les url des livres par catégorie.
    Elle prend en paramètre:
        - l'url de la page d'une catégorie
        - un tableau vide
        - le numéro de page  
    """
    
    #récuperation de l'url de la page courant et instanciation d'un objet "soup"
    reponse = requests.get(url_article)
    CONTENT_CATEGORY_PAGE = reponse.text

    soup = BeautifulSoup(CONTENT_CATEGORY_PAGE, "html.parser")

    # récupération des balises h3 et de leurs nombres par pages
    all_h3 = soup.find_all('h3')

    for h3 in all_h3:
        article_link = h3.a["href"]
        tab.append(article_link.replace("../../../", f"{URL_HOME}catalogue/"))
        print(f"Récupération de l'url du livre : {h3.a.text}.")
        
    # vérififcation d'un bouton next sur la page
    if not is_many_pages(soup):
        return tab
    elif is_many_pages(soup) and numero_page == 1:
        url_article = url_article.replace("index.html", f"page-{numero_page}.html")
        empty_tab = []
        return get_links_articles_for_one_categorie(url_article, empty_tab, numero_page+1)
    elif is_many_pages(soup):
        url_article = url_article.replace(f"page-{numero_page-1}.html", f"page-{numero_page}.html")
        return get_links_articles_for_one_categorie(url_article,tab, numero_page+1)



def get_links_articles_by_categories(dict_categories_and_links):
    dict_urls_by_categories = {}
    for categorie, link in dict_categories_and_links.items():
        list_urls_by_categories = []
        print(f"Récupération des urls des livres pour la catégorie {categorie}.")
        dict_urls_by_categories[categorie] = get_links_articles_for_one_categorie(link, list_urls_by_categories ,1)

    return dict_urls_by_categories


def image_download(url_image, article_name, image_folder):
    reponse = requests.get(url_image)
    
    if reponse.status_code == 200:
        article_name = re.sub('[^a-zA-Z]+', '', article_name).lower()
        salt = random.randint(0,100)
        if len(article_name) > 40:
            article_name = article_name[0:30]
        f = open(f'Categories/{image_folder}/{article_name}-{salt}.jpg', 'wb')
        f.write(reponse.content)
        f.close()
        print(f"Téléchargement de l'image du livre {article_name} terminée.")
        return True
    else:
        print(f"ERREUR : {reponse.status_code}")


def get_informations_book(book_url):
    """
    Cette fonction permet de récuperer les informations d'un livre à partir de son url.
    Elle prend en paramètre l'url du livre en question.
    Elle retourne ces informations dans un dictionnaire

    """
    requete = requests.get(book_url)
    requete.encoding = requete.apparent_encoding

    soup = BeautifulSoup(requete.text, "html.parser")

    # titre
    article_name = soup.h1.text

    # récuperation des notes
    block_article = soup('div', class_="product_main")
    for informations in block_article:
        rate = informations.select_one('p[class*="star-rating"]')['class'][1]

    match rate:
        case "One":
            rate = "1"
        case "Two":
            rate = "2"
        case "Three":
            rate = "3"
        case "Four":
            rate = "4"
        case "Five":
            rate = "5"
        case _:
            rate = "aucune note"

    # description
    if not soup.find("div", id="product_description"):
        description = "aucune description" 
    else:
        description = soup.find("div", id="product_description").find_next('p').text
    description = description.replace(",", "")

    # tableau infos
    informations_table = soup.find("table", class_="table table-striped")

    # image
    bloc_image = soup.find('div', id="product_gallery")
    image_url = bloc_image.div.div.div.img['src'].replace("../../", URL_HOME)

    # création dictionnaire
    dictionnary_informations_book = {}

    dictionnary_informations_book["Nom"]           = article_name
    dictionnary_informations_book["Note"]          = rate
    dictionnary_informations_book["Description"]   = strip_description
    for data in informations_table:
        header = data.get_text(",", True).split(",")
        if len(header) > 1:
            match header[0]:
                case "Price (excl. tax)":
                    price_exclude_tax = header[0].replace('Price' ,"Price (£) ")
                    dictionnary_informations_book[price_exclude_tax] = header[1].removeprefix("£")
                case "Price (incl. tax)":
                    price_include_tax = header[0].replace('Price' ,"Price (£) ")
                    dictionnary_informations_book[price_include_tax] = header[1].removeprefix("£")
                case "Tax":
                    tax = header[0] + " (£)"
                    dictionnary_informations_book[tax] = header[1].removeprefix("£")
                case "Availability":
                    number_book_in_stock = header[1].removeprefix("In stock (").removesuffix(" available)")
                    dictionnary_informations_book[header[0]] = number_book_in_stock
                case _:
                    dictionnary_informations_book[header[0]] = header[1]
    dictionnary_informations_book["Image"]         = image_url

    return dictionnary_informations_book



def save_data_by_categorie(dict_url:dict):
    os.makedirs(f"Categories/Images", exist_ok=True)
    for categorie, liste_url in dict_url.items():
        os.makedirs(f"Categories/Images/{categorie}", exist_ok=True)
        path_img = f"Images/{categorie}/"
        liste_infos_articles_categorie = []
        
        for url in liste_url:
            print(f"récupération infos par livres de la catégorie {categorie} en cours...")
            liste_infos_articles_categorie.append(get_informations_book(url))


        print(f"impression des données de la catégorie {categorie} en cours ...")
        fichier = openpyxl.load_workbook(f"Categories/{categorie}.xlsx")
        sheet = fichier.active

        row = 2
        
        for i in range(len(liste_infos_articles_categorie)):
            coll = 1
            for keys, values in liste_infos_articles_categorie[i].items():
                if keys == "Image":
                    url_img = liste_infos_articles_categorie[i]['Image']
                    name_img = liste_infos_articles_categorie[i]['Nom']
                    image_download(url_img, name_img,path_img)

                sheet.cell(1, coll).font = Font(size = 20, bold = True, vertAlign= "baseline")
                sheet.cell(1, coll).alignment = Alignment(horizontal="center")
                sheet.cell(1, coll).value = keys

                
                sheet.cell(row,coll).value = values
                sheet.cell(row,coll).alignment = Alignment(horizontal="left")

                if keys == "Description":
                    sheet.cell(row,coll).alignment = Alignment(horizontal="left")

                coll+=1
                
            row+=1

            # Gestion de la largeur des colonnes
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try: 
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                if column == "A" or column == "D" or column == "K":
                    adjusted_width = (max_length + 2)
                elif column == "C":
                    adjusted_width = 100
                else:
                    adjusted_width = (max_length + 2) * 1.8
                sheet.column_dimensions[column].width = adjusted_width
        



        print(f"l'impression des données pour la catégorie {categorie} a bien été effectuée.")
        fichier.save(f"Categories/{categorie}.xlsx")







